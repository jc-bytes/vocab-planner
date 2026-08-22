from copy import copy
from pathlib import Path
import sys

from openpyxl import load_workbook


path = Path(sys.argv[1])
criterion_count = int(sys.argv[2])
workbook = load_workbook(path)
sheet = workbook["Template"]

# The Classroom template contains five fully styled criterion blocks. Extend the
# fifth block's presentation to any additional criteria without changing values.
for criterion_index in range(5, criterion_count):
    destination_start = 3 + criterion_index * 5
    for offset in range(5):
        source_row = 23 + offset
        destination_row = destination_start + offset
        sheet.row_dimensions[destination_row].height = sheet.row_dimensions[source_row].height
        for column in range(1, 6):
            source = sheet.cell(source_row, column)
            destination = sheet.cell(destination_row, column)
            destination._style = copy(source._style)
            destination.number_format = source.number_format
            destination.protection = copy(source.protection)
            destination.alignment = copy(source.alignment)

workbook.save(path)
